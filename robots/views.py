from datetime import date, timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import Count
from io import BytesIO
from .models import Robot


def download_excel(request):
    today = date.today()
    week = today - timedelta(days=7)
    robots = Robot.objects.filter(created__range=(week, today)).values('model', 'version').annotate(week_count=Count('id'))

    wb = Workbook()
    wb.remove(wb.active)

    for model in robots.values('model').distinct():
        model_name = model['model']
        sheet = wb.create_sheet(title=model_name)
        sheet.append(['Модель', 'Версия', 'Количество за неделю'])
        model_data = robots.filter(model=model_name)
        for robot in model_data:
            sheet.append([robot['model'], robot['version'], robot['week_count']])
        for col in range(1, 4):
            column = get_column_letter(col)
            max_length = 0
            for row in sheet.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    response = HttpResponse(excel_file, content_type='text/csv; charset=utf-8-sig' )
    response['Content-Disposition'] = 'attachment; filename="weekly_report.csv"'
    return response